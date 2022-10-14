#/user/bin/python3

import openpyxl
import math

wb = openpyxl.load_workbook("student.xlsx")
ws = wb['Sheet1']


# total
data = []
row_id = 1
for row in ws:
	if row_id != 1:
		sum_v = ws.cell(row = row_id, column = 3).value * 0.3
		sum_v += ws.cell(row = row_id, column = 4).value * 0.35
		sum_v += ws.cell(row = row_id, column = 5).value * 0.34
		sum_v += ws.cell(row = row_id, column = 6).value
		ws.cell(row = row_id, column = 7).value = sum_v
		data.append([row_id, sum_v, 0])
	row_id += 1


# grade 
data.sort(key=lambda x:x[1], reverse=True) 
same = 0
rank = 1
for i in range(len(data) - 1):
    if data[i][1] == data[i+1][1]:
        same += 1
    else:
        while (1): 
            data[i - same][2] = rank
            if (same == 0):
                break
            same -= 1
    rank += 1

while (1):
	data[len(data) - 1 - same][2] = rank
	if (same == 0):
		break
	same -= 1

n = len(data)
A_plus = math.floor(n * 0.3 * 0.5)
A_zero = math.floor(n * 0.3)
B_plus = math.floor(n * 0.7 * 0.5)
B_zero = math.floor(n * 0.7)
C_plus = B_zero + math.floor((n - B_zero) * 0.5)

row_id = 0
for row in ws:
    rank = data[row_id][2]
    if rank <= A_plus:
        ws.cell(row = data[row_id][0], column = 8).value = 'A+'
    elif rank <= A_zero:
        ws.cell(row = data[row_id][0], column = 8).value = 'A0'
    elif rank <= B_plus:
        ws.cell(row = data[row_id][0], column = 8).value = 'B+'
    elif rank <= B_zero:
        ws.cell(row = data[row_id][0], column = 8).value = 'B0'
    elif rank <= C_plus:
        ws.cell(row = data[row_id][0], column = 8).value = 'C+'
    else:
        ws.cell(row = data[row_id][0], column = 8).value = 'C0'
    row_id += 1
    if (row_id == n):
        break

wb.save("student.xlsx")

